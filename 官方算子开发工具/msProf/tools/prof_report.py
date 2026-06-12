"""
prof_report.py — msprof op 采集结果 Excel 报告生成器

用法：
  python prof_report.py <OPPROF目录>              # 单组报告
  python prof_report.py <目录A> <目录B>           # 两组对比报告

输出：报告放在第一个输入目录下，文件名 prof_report.xlsx

依赖：pip install openpyxl
"""

import sys, os, csv, glob
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers as xlnums)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError:
    print("缺少 openpyxl，请先运行：pip install openpyxl")
    sys.exit(1)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ── 样式常量 ──────────────────────────────────────────────────────────────────

C_HEADER  = "FF2C3E50"   # 深蓝黑，表头背景
C_WHITE   = "FFFFFFFF"
C_LIGHT   = "FFF5F6FA"   # 交替行浅色
C_MTE2    = "FFB3D4FF"   # 蓝
C_MTE3    = "FFFFD580"   # 黄
C_SCALAR  = "FFD4EDDA"   # 绿
C_SLOW    = "FFFFE0E0"   # 粉红，最慢 block 高亮
C_WARN    = "FFFF6B6B"   # 红，告警文字

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="FFBDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(C_HEADER)
    c.font = Font(color=C_WHITE, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value, bg=None, bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg or C_WHITE)
    c.font = Font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

# ── CSV 解析 ──────────────────────────────────────────────────────────────────

def resolve_csv(directory, filename):
    direct = os.path.join(directory, filename)
    if os.path.exists(direct):
        return direct
    matches = glob.glob(os.path.join(directory, f"*_{filename}"))
    return matches[0] if matches else direct

def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def fval(row, key):
    v = row.get(key, "").strip()
    if not v or v.upper() == "NA":
        return None
    try:
        return float(v)
    except ValueError:
        return None

def avg_col(rows, key):
    vals = [fval(r, key) for r in rows]
    vals = [v for v in vals if v is not None]
    return sum(vals) / len(vals) if vals else None

def load_data(opprof_dir):
    basic_rows = read_csv(resolve_csv(opprof_dir, "OpBasicInfo.csv"))
    info = basic_rows[0] if basic_rows else {}
    pipe_rows = read_csv(resolve_csv(opprof_dir, "PipeUtilization.csv"))
    pipe_rows = [r for r in pipe_rows if fval(r, "aic_time(us)") is not None]
    return info, pipe_rows

PIPE_DEFS = [
    ("MTE2 (GM→UB)", "aic_mte2_time(us)", "aic_mte2_ratio", "aic_mte2_active_bw(GB/s)", C_MTE2),
    ("MTE3 (UB→GM)", "aic_mte3_time(us)", "aic_mte3_ratio", "aic_mte3_active_bw(GB/s)", C_MTE3),
    ("Scalar",       "aic_scalar_time(us)","aic_scalar_ratio", None,                     C_SCALAR),
    ("Cube",         "aic_cube_time(us)",  "aic_cube_ratio",  None,                      C_LIGHT),
    ("MTE1",         "aic_mte1_time(us)",  "aic_mte1_ratio",  "aic_mte1_active_bw(GB/s)",C_LIGHT),
    ("Fixpipe",      "aic_fixpipe_time(us)","aic_fixpipe_ratio", None,                   C_LIGHT),
]

# ── Sheet 1：汇总 ─────────────────────────────────────────────────────────────

def write_summary(wb, info, pipe_rows, label=""):
    ws = wb.active
    ws.title = "汇总"
    ws.row_dimensions[1].height = 30

    # 标题
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"msprof op 性能报告  {('— ' + label) if label else ''}  ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
    t.font = Font(size=13, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # 基本信息
    r = 3
    for key, val in [
        ("Op Name",      info.get("Op Name","").split("__kernel")[0][-60:]),
        ("Op Type",      info.get("Op Type","")),
        ("Block Dim",    info.get("Block Dim","")),
        ("Task Duration","({:.3f} μs  ← 含驱动开销)".format(fval(info, "Task Duration(us)") or 0)),
        ("Rated Freq",   info.get("Rated Freq","") + " MHz"),
    ]:
        ws.cell(row=r, column=1, value=key).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1

    # PIPE 均值表
    r += 1
    if not pipe_rows:
        return

    aic_times = [fval(row, "aic_time(us)") for row in pipe_rows]
    slowest_i = aic_times.index(max(aic_times))
    typical = [row for i, row in enumerate(pipe_rows) if i != slowest_i] or pipe_rows

    hdr_cell(ws, r, 1, "PIPE 单元",    18)
    hdr_cell(ws, r, 2, "典型时间(μs)", 15)
    hdr_cell(ws, r, 3, "占比(%)",      10)
    hdr_cell(ws, r, 4, "活跃带宽(GB/s)", 16)
    r += 1

    bottleneck = None
    max_ratio = 0
    for name, t_key, r_key, bw_key, color in PIPE_DEFS:
        t   = avg_col(typical, t_key)
        rat = avg_col(typical, r_key)
        bw  = avg_col(typical, bw_key) if bw_key else None
        if rat and rat > max_ratio:
            max_ratio = rat
            bottleneck = name
        data_cell(ws, r, 1, name,                   bg=color, align="left")
        data_cell(ws, r, 2, round(t,   3) if t   is not None else "N/A", bg=color, fmt="0.000")
        data_cell(ws, r, 3, round(rat*100, 1) if rat is not None else "N/A", bg=color, fmt="0.0")
        data_cell(ws, r, 4, round(bw,  2) if bw  is not None else "N/A", bg=color, fmt="0.00")
        r += 1

    # Per-block 时间
    r += 1
    min_t, max_t = min(aic_times), max(aic_times)
    mean_t = sum(aic_times) / len(aic_times)
    ws.cell(row=r, column=1, value="Per-block AIC 时间").font = Font(bold=True)
    r += 1
    for label_s, val in [("最快", f"{min_t:.3f} μs  (block {aic_times.index(min_t)})"),
                          ("最慢", f"{max_t:.3f} μs  (block {slowest_i})"
                                   + ("  ★ 不均衡" if max_t > min_t * 1.3 else "")),
                          ("平均", f"{mean_t:.3f} μs")]:
        ws.cell(row=r, column=2, value=label_s)
        c = ws.cell(row=r, column=3, value=val)
        if "不均衡" in (val or ""):
            c.font = Font(color=C_WARN, bold=True)
        r += 1

    # 结论
    r += 1
    concl = ws.cell(row=r, column=1,
                    value=f"★ 瓶颈：{bottleneck} 主导 → {'Memory-bound' if 'MTE' in (bottleneck or '') else 'Compute-bound / 其他'}")
    concl.font = Font(bold=True, color=C_WARN)
    ws.merge_cells(f"A{r}:F{r}")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16

# ── Sheet 2：Per-block 明细 ───────────────────────────────────────────────────

def write_block_detail(wb, pipe_rows, sheet_name="Block明细"):
    ws = wb.create_sheet(sheet_name)

    if not pipe_rows:
        ws["A1"] = "无数据"
        return

    aic_times = [fval(row, "aic_time(us)") for row in pipe_rows]
    slowest_i = aic_times.index(max(aic_times))

    # 表头
    headers = ["Block", "AIC时间(μs)", "cycles",
               "MTE2时间", "MTE2%", "MTE3时间", "MTE3%", "MTE3带宽(GB/s)",
               "Scalar%", "Cube%", "icache_miss%"]
    widths  = [8, 13, 10, 11, 8, 11, 8, 15, 10, 8, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        hdr_cell(ws, 1, c, h, w)

    for i, row in enumerate(pipe_rows):
        r = i + 2
        is_slow = (i == slowest_i)
        bg = C_SLOW if is_slow else (C_LIGHT if i % 2 == 0 else C_WHITE)
        bold = is_slow

        def dc(col, val, fmt=None):
            data_cell(ws, r, col, val, bg=bg, bold=bold, fmt=fmt)

        dc(1,  i)
        dc(2,  fval(row, "aic_time(us)"),        "0.000")
        dc(3,  fval(row, "aic_total_cycles"),     "0")
        dc(4,  fval(row, "aic_mte2_time(us)"),    "0.000")
        dc(5,  round((fval(row,"aic_mte2_ratio") or 0)*100, 1), "0.0")
        dc(6,  fval(row, "aic_mte3_time(us)"),    "0.000")
        dc(7,  round((fval(row,"aic_mte3_ratio") or 0)*100, 1), "0.0")
        dc(8,  fval(row, "aic_mte3_active_bw(GB/s)"), "0.00")
        dc(9,  round((fval(row,"aic_scalar_ratio") or 0)*100, 1), "0.0")
        dc(10, round((fval(row,"aic_cube_ratio")  or 0)*100, 1), "0.0")
        dc(11, round((fval(row,"aic_icache_miss_rate") or 0)*100, 2), "0.00")

    if slowest_i + 2 <= len(pipe_rows) + 1:
        note_r = len(pipe_rows) + 3
        ws.merge_cells(f"A{note_r}:K{note_r}")
        ws[f"A{note_r}"] = f"粉色行 = 最慢 block（block {slowest_i}）"
        ws[f"A{note_r}"].font = Font(italic=True, color="FF888888")

    ws.freeze_panes = "A2"

# ── Sheet 3：Block 均衡柱状图 ─────────────────────────────────────────────────

def write_block_chart(wb, pipe_rows, label=""):
    ws = wb.create_sheet("Block均衡图")

    if not pipe_rows:
        ws["A1"] = "无数据"
        return

    # 写数据供图表引用
    ws["A1"] = "Block"
    ws["B1"] = "AIC时间(μs)"
    ws["C1"] = "MTE2时间(μs)"
    ws["D1"] = "MTE3时间(μs)"

    for i, row in enumerate(pipe_rows):
        r = i + 2
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=fval(row, "aic_time(us)"))
        ws.cell(row=r, column=3, value=fval(row, "aic_mte2_time(us)"))
        ws.cell(row=r, column=4, value=fval(row, "aic_mte3_time(us)"))

    n = len(pipe_rows)

    # AIC 总时间柱状图
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = f"Per-block AIC 总时间  {label}"
    chart1.y_axis.title = "μs"
    chart1.x_axis.title = "Block ID"
    chart1.shape = 4
    chart1.width = 20
    chart1.height = 12

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=n + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.solidFill = "4472C4"
    ws.add_chart(chart1, "F1")

    # MTE2 + MTE3 堆叠柱状图
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = f"MTE2 / MTE3 时间分布  {label}"
    chart2.y_axis.title = "μs"
    chart2.x_axis.title = "Block ID"
    chart2.width = 20
    chart2.height = 12

    for col, color in [(3, "5B9BD5"), (4, "ED7D31")]:
        ref = Reference(ws, min_col=col, max_col=col, min_row=1, max_row=n + 1)
        chart2.add_data(ref, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.series[0].graphicalProperties.solidFill = "5B9BD5"
    chart2.series[1].graphicalProperties.solidFill = "ED7D31"
    ws.add_chart(chart2, "F22")

# ── Sheet 4：对比 ─────────────────────────────────────────────────────────────

def write_compare(wb, data_list, labels):
    ws = wb.create_sheet("对比")

    infos     = [d[0] for d in data_list]
    pipe_list = [d[1] for d in data_list]

    # 对比表头
    hdr_cell(ws, 1, 1, "指标", 22)
    for i, lbl in enumerate(labels):
        hdr_cell(ws, 1, i + 2, lbl, 18)
    hdr_cell(ws, 1, len(labels) + 2, "变化", 14)

    metrics = [
        ("Task Duration (μs)", lambda info, _: fval(info, "Task Duration(us)")),
        ("Block Dim",          lambda info, _: info.get("Block Dim", "")),
    ]

    def avg_pipe(pipe_rows, key):
        if not pipe_rows:
            return None
        times = [fval(r, "aic_time(us)") for r in pipe_rows]
        slowest_i = times.index(max(times))
        typical = [r for i, r in enumerate(pipe_rows) if i != slowest_i] or pipe_rows
        return avg_col(typical, key)

    pipe_metrics = [
        ("Mean block AIC (μs)", "aic_time(us)"),
        ("MTE2 占比 (%)",       "aic_mte2_ratio"),
        ("MTE3 占比 (%)",       "aic_mte3_ratio"),
        ("MTE3 带宽 (GB/s)",    "aic_mte3_active_bw(GB/s)"),
    ]

    r = 2
    for label_m, fn in metrics:
        vals = [fn(info, pipe) for info, pipe in zip(infos, pipe_list)]
        data_cell(ws, r, 1, label_m, align="left")
        for c, v in enumerate(vals, 2):
            data_cell(ws, r, c, v, bg=C_LIGHT if r % 2 == 0 else C_WHITE)
        data_cell(ws, r, len(labels) + 2, "")
        r += 1

    for label_m, key in pipe_metrics:
        scale = 100 if "占比" in label_m else 1
        vals = [avg_pipe(pipe_rows, key) for pipe_rows in pipe_list]
        vals_scaled = [round(v * scale, 3) if v is not None else None for v in vals]
        data_cell(ws, r, 1, label_m, align="left")
        for c, v in enumerate(vals_scaled, 2):
            data_cell(ws, r, c, v, bg=C_LIGHT if r % 2 == 0 else C_WHITE,
                      fmt="0.00" if v is not None and isinstance(v, float) else None)

        # 变化列
        change = ""
        if len(vals_scaled) >= 2 and vals_scaled[0] and vals_scaled[1]:
            ratio = vals_scaled[1] / vals_scaled[0]
            if ratio < 0.85:
                change = f"↓ {1/ratio:.1f}×"
            elif ratio > 1.15:
                change = f"↑ {ratio:.1f}×"
            else:
                change = "≈ 相当"
        c_cell = data_cell(ws, r, len(labels) + 2, change)
        if "↓" in change:
            c_cell.font = Font(color="FF008000", bold=True)
        elif "↑" in change:
            c_cell.font = Font(color=C_WARN, bold=True)
        r += 1

    # 对比柱状图（Task Duration）
    chart = BarChart()
    chart.type = "col"
    chart.title = "Task Duration 对比 (μs)"
    chart.y_axis.title = "μs"
    chart.width = 16
    chart.height = 10

    # 写一个小数据区供图表用
    data_start_r = r + 2
    ws.cell(row=data_start_r, column=1, value="版本")
    ws.cell(row=data_start_r, column=2, value="Task Duration(μs)")
    for i, (info, lbl) in enumerate(zip(infos, labels)):
        ws.cell(row=data_start_r + 1 + i, column=1, value=lbl)
        ws.cell(row=data_start_r + 1 + i, column=2, value=fval(info, "Task Duration(us)"))

    n = len(infos)
    data_ref = Reference(ws, min_col=2, max_col=2,
                         min_row=data_start_r, max_row=data_start_r + n)
    cats_ref = Reference(ws, min_col=1,
                         min_row=data_start_r + 1, max_row=data_start_r + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E2")

# ── 主流程 ────────────────────────────────────────────────────────────────────

def has_prof_csv(directory):
    """目录里是否有 msProf 采集结果 CSV（直接命中或带前缀命中均可）。"""
    for name in ("OpBasicInfo.csv", "PipeUtilization.csv"):
        if os.path.exists(resolve_csv(directory, name)):
            return True
    return False


def main():
    args = sys.argv[1:]
    if not args:
        d = sorted(glob.glob("OPPROF_*"))
        if not d:
            print("未找到 OPPROF_* 目录，请指定路径")
            sys.exit(1)
        args = [d[-1]]

    # 输入校验：避免对着空目录默默生成空表
    for a in args:
        if not os.path.isdir(a):
            print(f"ERROR: 目录不存在：{a}")
            sys.exit(1)
        if not has_prof_csv(a):
            print(f"ERROR: 该目录里没有 OpBasicInfo.csv / PipeUtilization.csv，"
                  f"不是 msProf 采集结果目录：{a}")
            print("  请把“msProf 输出目录”指向真正包含这些 CSV 的采集结果目录"
                  "（通常是 msprof op 采集生成的 OPPROF_* 目录）。")
            sys.exit(1)

    labels = [os.path.basename(a.rstrip("/\\")) for a in args]
    data_list = [load_data(a) for a in args]

    wb = openpyxl.Workbook()

    if len(args) == 1:
        info, pipe_rows = data_list[0]
        write_summary(wb, info, pipe_rows)
        write_block_detail(wb, pipe_rows)
        write_block_chart(wb, pipe_rows, labels[0])
    else:
        # 多组：汇总用第一组，再写对比页
        info, pipe_rows = data_list[0]
        write_summary(wb, info, pipe_rows, labels[0])
        for i, (inf, pr) in enumerate(data_list):
            write_block_detail_named(wb, pr, labels[i])
            write_block_chart(wb, pr, labels[i])
        write_compare(wb, data_list, labels)

    out_path = os.path.join(args[0], "prof_report.xlsx")
    wb.save(out_path)
    print(f"已生成：{out_path}")
    print("  Sheet: 汇总 / Block明细 / Block均衡图" +
          (" / 对比" if len(args) > 1 else ""))
    # 工具箱插件协议：成功时 stdout 最后一行输出生成物路径
    print(f"OUTPUT={out_path}")


def write_block_detail_named(wb, pipe_rows, label):
    """多组模式下给每组单独建一个明细 sheet（复用 write_block_detail，仅改 sheet 名）"""
    write_block_detail(wb, pipe_rows, sheet_name=f"Block明细-{label}"[:31])


if __name__ == "__main__":
    main()
