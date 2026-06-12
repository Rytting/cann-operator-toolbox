"""
仿真流水图 Excel 报告生成脚本
用法: python trace_report.py <dump2trace_core0.json> [output.xlsx]

依赖: pip install openpyxl
"""
import json, sys, os, re
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── 单元颜色（和 Chrome tracing 图里接近）────────────────────────────
UNIT_COLORS = {
    'RVECSU': 'FF4CAF50',   # 绿
    'RVECEX': 'FFCD853F',   # 棕
    'RVECLD': 'FF4169E1',   # 蓝
    'RVECST': 'FFCD5C5C',   # 红
    'PUSHQ':  'FF90EE90',   # 浅绿
}
UNIT_ORDER = ['RVECLD', 'RVECSU', 'RVECEX', 'RVECST', 'PUSHQ']
HEADER_FILL = PatternFill('solid', fgColor='FF2C3E50')
HEADER_FONT = Font(color='FFFFFFFF', bold=True)
ALT_FILL    = PatternFill('solid', fgColor='FFF2F3F4')

def thin_border():
    s = Side(style='thin', color='FFBDC3C7')
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def parse_instr_name(detail):
    """从 detail 字符串提取指令名，如 '(ID: 004567) RV_PLT ...' → 'RV_PLT'"""
    m = re.search(r'\)\s+(\S+)', detail)
    return m.group(1) if m else '?'

def parse_instr_id(detail):
    m = re.search(r'ID:\s*(\d+)', detail)
    return int(m.group(1)) if m else 0

def load_events(path):
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    events = [e for e in data['traceEvents'] if e.get('ph') == 'X']
    for e in events:
        detail = e['args'].get('detail', '')
        e['instr_name'] = parse_instr_name(detail)
        e['instr_id']   = parse_instr_id(detail)
    return events

def compute_stats(events):
    t_start = min(e['ts'] for e in events)
    t_end   = max(e['ts'] + e['dur'] for e in events)
    span    = t_end - t_start

    by_unit = defaultdict(list)
    for e in events:
        by_unit[e['tid']].append(e)

    rows = []
    for unit in UNIT_ORDER:
        if unit not in by_unit:
            continue
        evts = sorted(by_unit[unit], key=lambda e: e['ts'])
        active = sum(e['dur'] for e in evts)
        longest = max(evts, key=lambda e: e['dur'])
        # 最大空闲间隔
        gaps = [evts[i]['ts'] - (evts[i-1]['ts'] + evts[i-1]['dur'])
                for i in range(1, len(evts))]
        gaps = [g for g in gaps if g > 0]
        rows.append({
            'unit':        unit,
            'count':       len(evts),
            'active_us':   active,
            'util_pct':    active / span * 100,
            'max_gap_us':  max(gaps) if gaps else 0,
            'longest_us':  longest['dur'],
            'longest_instr': longest['instr_name'],
        })

    bottleneck = max(rows, key=lambda r: r['active_us'])['unit']
    return span, rows, bottleneck, by_unit

def write_summary_sheet(wb, span_us, stat_rows, bottleneck, json_name):
    ws = wb.active
    ws.title = '汇总'

    # 标题行
    ws.merge_cells('A1:G1')
    ws['A1'] = f'算子仿真流水图统计 — {json_name}'
    ws['A1'].font = Font(size=13, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f'Kernel 总时间跨度：{span_us/1000:.3f} ms（{span_us:,} μs）'
    ws['A2'].font = Font(italic=True, color='FF555555')
    ws['A3'] = f'生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, color='FF555555')

    # 表头
    headers = ['功能单元', '指令条数', '活跃时间 (μs)', '利用率 (%)',
               '最大空闲间隔 (μs)', '最长单指令 (μs)', '最长指令类型']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    # 数据行
    for r, row in enumerate(stat_rows, 6):
        vals = [row['unit'], row['count'], row['active_us'],
                round(row['util_pct'], 2), row['max_gap_us'],
                row['longest_us'], row['longest_instr']]
        color = UNIT_COLORS.get(row['unit'], 'FFFFFFFF')
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
            if row['unit'] == bottleneck:
                cell.font = Font(bold=True)

    # 瓶颈说明
    note_row = 6 + len(stat_rows) + 1
    ws.merge_cells(f'A{note_row}:G{note_row}')
    ws[f'A{note_row}'] = f'★ 瓶颈单元：{bottleneck}（活跃时间最长，加粗标注）'
    ws[f'A{note_row}'].font = Font(bold=True, color='FFCC0000')

    note_row += 1
    ws.merge_cells(f'A{note_row}:G{note_row}')
    ws[f'A{note_row}'] = ('注：利用率 = 该单元活跃时间 / Kernel 总时间跨度。'
                          '若双缓冲循环运行多次，总时间跨度偏大，利用率偏低属正常，'
                          '请以"活跃时间"和"最长单指令"为主要参考。')
    ws[f'A{note_row}'].font = Font(italic=True, color='FF777777')
    ws.row_dimensions[note_row].height = 30
    ws[f'A{note_row}'].alignment = Alignment(wrap_text=True)

    # 列宽
    for col, w in zip(range(1, 8), [12, 10, 16, 12, 18, 16, 16]):
        set_col_width(ws, col, w)

def write_detail_sheet(wb, events, by_unit):
    ws = wb.create_sheet('指令明细')

    headers = ['序号', '功能单元', '指令ID', '指令类型',
               '开始时间 (μs)', '持续时间 (μs)', '内存地址']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    # 按开始时间排序
    sorted_events = sorted(events, key=lambda e: e['ts'])
    t0 = sorted_events[0]['ts']

    for r, e in enumerate(sorted_events, 2):
        color = UNIT_COLORS.get(e['tid'], 'FFFFFFFF')
        fill = PatternFill('solid', fgColor=color) if r % 2 == 0 else ALT_FILL
        vals = [r - 1, e['tid'], e['instr_id'], e['instr_name'],
                e['ts'] - t0, e['dur'],
                e['args'].get('addr', '')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()

    # 列宽
    for col, w in zip(range(1, 8), [6, 10, 10, 14, 16, 16, 14]):
        set_col_width(ws, col, w)

    ws.freeze_panes = 'A2'

def write_topN_sheet(wb, events, n=20):
    ws = wb.create_sheet(f'耗时 Top{n}')

    headers = ['排名', '功能单元', '指令类型', '持续时间 (μs)', '开始时间 (μs)', '内存地址']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    top = sorted(events, key=lambda e: e['dur'], reverse=True)[:n]
    t0 = min(e['ts'] for e in events)

    for r, e in enumerate(top, 2):
        color = UNIT_COLORS.get(e['tid'], 'FFFFFFFF')
        vals = [r - 1, e['tid'], e['instr_name'],
                e['dur'], e['ts'] - t0,
                e['args'].get('addr', '')]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()

    for col, w in zip(range(1, 7), [6, 10, 14, 16, 16, 14]):
        set_col_width(ws, col, w)

def main():
    json_path = sys.argv[1] if len(sys.argv) > 1 else 'dump2trace_core0.json'
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else json_path.replace('.json', '_report.xlsx')

    if not os.path.exists(json_path):
        print(f'ERROR: 找不到输入文件: {json_path}')
        sys.exit(1)

    print(f'读取: {json_path}')
    events = load_events(json_path)
    span_us, stat_rows, bottleneck, by_unit = compute_stats(events)

    wb = openpyxl.Workbook()
    write_summary_sheet(wb, span_us, stat_rows, bottleneck, os.path.basename(json_path))
    write_detail_sheet(wb, events, by_unit)
    write_topN_sheet(wb, events, n=20)

    wb.save(xlsx_path)
    print(f'已生成: {xlsx_path}')
    print(f'  - 汇总：各单元活跃时间、利用率、最大停顿')
    print(f'  - 指令明细：全部 {len(events)} 条指令按时间排序')
    print(f'  - 耗时 Top20：持续时间最长的 20 条指令')
    # 工具箱插件协议：成功时 stdout 最后一行输出生成物路径
    print(f'OUTPUT={xlsx_path}')

if __name__ == '__main__':
    main()
